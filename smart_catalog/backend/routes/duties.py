"""职能管理路由 — 上传解析、铁律一（唯一性校验）、匹配执行"""

import json
import os
import io
import uuid
from datetime import datetime
from flask import Blueprint, request, jsonify, current_app, send_file
import pandas as pd
import openpyxl
from models import db, Duty, Line, LineDutyDept, Catalog, MatchResult
from utils import bad_request, paginate, jsonify_paginated

duties_bp = Blueprint('duties', __name__, url_prefix='/api/duties')


def _parse_duty_upload(file_storage):
    """解析职能清单 Excel/CSV"""
    _, ext = os.path.splitext(file_storage.filename.lower())
    try:
        if ext in ('.xlsx', '.xls'):
            df = pd.read_excel(file_storage, dtype=str)
        elif ext == '.csv':
            df = pd.read_csv(file_storage, dtype=str, encoding='utf-8')
        else:
            return None, '不支持的文件格式'
    except Exception as e:
        return None, f'文件解析失败：{str(e)}'

    if df.empty:
        return None, '文件为空'

    col_map = {}
    for col in df.columns:
        cl = col.strip().lower().replace(' ', '').replace('　', '')
        if cl in ('部门名称', '部门', 'department', 'dept', '单位'):
            col_map['department'] = col
        elif cl in ('内设机构', '内设', '机构', 'inner', 'org'):
            col_map['inner_org'] = col
        elif cl in ('职能名称', '职能', '职责', 'duty', 'duty_name', 'function'):
            col_map['duty_name'] = col

    if 'duty_name' not in col_map:
        return None, '未找到「职能名称」列'

    rows = []
    for _, row in df.iterrows():
        item = {}
        for key, col in col_map.items():
            val = str(row[col]).strip() if pd.notna(row[col]) else ''
            item[key] = val
        rows.append(item)

    return rows, None


def _validate_duty_unique(department):
    """
    铁律一：校验职责部门在「部门-条线映射表」中的映射数量
    返回 (is_valid, error_msg_or_none)
    """
    mappings = LineDutyDept.query.filter_by(department=department).all()
    count = len(mappings)
    if count == 0:
        return False, f'部门「{department}」在映射表中不存在（0条映射）'
    elif count >= 2:
        return False, f'部门「{department}」存在 {count} 条映射（需唯一，当前冲突）'
    return True, None


@duties_bp.route('/parse', methods=['POST'])
def parse_upload():
    """上传解析职能清单 — 同时执行铁律一校验"""
    if 'file' not in request.files:
        return bad_request('请上传文件')
    file = request.files['file']
    if not file.filename:
        return bad_request('文件名为空')

    rows, err = _parse_duty_upload(file)
    if err:
        return bad_request(err)

    # 对每行部门做铁律一校验，同时查找映射条线
    violations = []
    for row in rows:
        dept = row.get('department', '')
        if dept:
            valid, error_msg = _validate_duty_unique(dept)
            row['_valid'] = valid
            row['_error'] = None if valid else error_msg
            if valid:
                mapping = LineDutyDept.query.filter_by(department=dept).first()
                row['line_id'] = mapping.line_id
                row['line_name'] = mapping.line.name if mapping.line else None
            else:
                violations.append(error_msg)
                row['line_id'] = None
                row['line_name'] = None
        else:
            row['_valid'] = False
            row['_error'] = '部门名称为空，无法映射'
            row['line_id'] = None
            row['line_name'] = None
            violations.append('存在部门名称为空的行')

    return jsonify({
        'rows': rows,
        'total': len(rows),
        'violations': violations,
        'has_violations': len(violations) > 0
    })


@duties_bp.route('/save', methods=['POST'])
def save_duties():
    """保存职能数据（仅保存校验通过的）"""
    data = request.get_json(silent=True) or {}
    rows = data.get('rows', [])
    if not rows:
        return bad_request('无数据可保存')

    # 检查是否还有未处理的违规
    violations = [r.get('_error') for r in rows if r.get('_valid') is False and r.get('_error')]
    if violations:
        return bad_request('存在未处理的违规项：' + '；'.join(violations))

    Duty.query.delete()
    for row in rows:
        duty = Duty(
            department=row.get('department', '').strip(),
            inner_org=row.get('inner_org', '').strip(),
            duty_name=row.get('duty_name', '').strip(),
            line_id=row.get('line_id')
        )
        db.session.add(duty)
    db.session.commit()
    return jsonify({'message': f'成功保存 {len(rows)} 条职能', 'total': len(rows)})


@duties_bp.route('', methods=['GET'])
def list_duties():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 1000, type=int)
    query = Duty.query.order_by(Duty.department, Duty.inner_org)
    items, total, page, per_page, total_pages = paginate(query, page, per_page, default_per_page=1000)
    return jsonify_paginated([{
        'id': d.id,
        'department': d.department,
        'inner_org': d.inner_org,
        'duty_name': d.duty_name,
        'line_id': d.line_id,
        'line_name': d.line.name if d.line else None
    } for d in items], total, page, per_page, total_pages)


@duties_bp.route('/<int:duty_id>', methods=['DELETE'])
def delete_duty(duty_id):
    duty = db.session.get(Duty, duty_id)
    if not duty:
        return bad_request('职能不存在', 404)
    db.session.delete(duty)
    db.session.commit()
    return jsonify({'message': '删除成功'})


# ---- 智能匹配 ----

def _call_external_api(catalog_items, duty_items):
    """
    调用职能匹配智能体（多文件上传模式）
    入参：目录项列表 [{name, fields, ...}], 职能列表 [{name, id, ...}]
    返回：{catalog_name: duty_name, ...}
    """
    mode = os.getenv('MATCH_MODE', 'api')

    # 本地模式：随机匹配（开发/演示用）
    if mode == 'local':
        result = {}
        for c in catalog_items:
            if duty_items:
                import random
                d = random.choice(duty_items)
                result[c['name']] = d['name']
            else:
                result[c['name']] = ''
        return result

    # API 模式
    import requests
    from io import BytesIO

    api_url = os.getenv('MATCH_API_URL',
        'http://2.142.92.117:30486/scene_gateway/ad441c290aef406fa9350344f513461c')
    auth_token = os.getenv('MATCH_AUTH_TOKEN',
        'f2fa523939374d8392c9eb50a6dcb245')
    scene_key = os.getenv('MATCH_SCENE_KEY',
        'ad441c290aef406fa9350344f513461c')

    # 构建临时 xlsx：职能文件
    duty_io = BytesIO()
    dwb = openpyxl.Workbook()
    dws = dwb.active
    dws.title = '职能'
    dws.append(['部门名称', '内设机构', '职能名称'])
    # 批量查询 duty 数据
    duty_ids = [d['id'] for d in duty_items]
    duty_records = {d.id: d for d in Duty.query.filter(Duty.id.in_(duty_ids)).all()}
    for d_item in duty_items:
        duty = duty_records.get(d_item['id'])
        if duty:
            dws.append([duty.department, duty.inner_org, duty.duty_name])
    dwb.save(duty_io)
    duty_io.seek(0)

    # 构建临时 xlsx：目录文件
    catalog_io = BytesIO()
    cwb = openpyxl.Workbook()
    cws = cwb.active
    cws.title = '目录'
    cws.append(['目录名称', '来源部门', '字段项'])
    for c in catalog_items:
        cws.append([c['name'], '', ', '.join(c.get('fields', []))])
    cwb.save(catalog_io)
    catalog_io.seek(0)

    request_id = datetime.now().strftime('%Y%m%d%H%M%S') + uuid.uuid4().hex[:8]

    files = {
        'zhineng': ('zhineng.xlsx', duty_io,
                     'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
        'mulu': ('mulu.xlsx', catalog_io,
                  'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
    }
    form_data = {
        'keyword': '职能匹配',
        'requestId': request_id,
        'sceneKey': scene_key,
        'dialogId': '',
    }
    headers = {'AuthToken': auth_token}

    try:
        resp = requests.post(api_url, headers=headers, data=form_data,
                              files=files, timeout=120)
        resp.raise_for_status()
        body = resp.json()
        raw_items = body.get('data', [])

        matched = {}
        for entry in raw_items:
            catalog_name = ''
            matched_org = ''
            for line in entry.split('\n'):
                line = line.strip()
                if line.startswith('【数据目录名称】') and '】：' in line:
                    catalog_name = line.split('】：', 1)[1].strip()
                elif line.startswith('【匹配内设机构】') and '】：' in line:
                    matched_org = line.split('】：', 1)[1].strip()
                    if '、' in matched_org:
                        matched_org = matched_org.split('、')[0].strip()

            if catalog_name and matched_org:
                # 按内设机构匹配 duty
                found = None
                other = None
                for duty in duty_records.values():
                    if duty.inner_org == matched_org or duty.department == matched_org:
                        found = duty
                        break
                    if not other and (matched_org in duty.inner_org or matched_org in duty.department):
                        other = duty
                duty = found or other
                matched[catalog_name] = duty.duty_name if duty else ''
            elif catalog_name:
                matched[catalog_name] = ''

        return matched

    except Exception as e:
        current_app.logger.error(f'API 调用失败：{str(e)}')
        raise RuntimeError(f'外部匹配接口调用失败：{str(e)}，请检查网络或接口配置')


@duties_bp.route('/match', methods=['POST'])
def match_duties():
    """
    铁律三：匹配执行 — 先分组，后分批
    1. 再次校验所有职责部门唯一性
    2. 按条线分组
    3. 每个条线下构造目录清单和职责清单调用外部接口
    """
    data = request.get_json(silent=True) or {}
    duty_ids = data.get('duty_ids', [])

    if duty_ids:
        duties = Duty.query.filter(Duty.id.in_(duty_ids)).all()
    else:
        duties = Duty.query.all()

    if not duties:
        return bad_request('没有待匹配的职能数据')

    # 铁律一：再次校验每个职责部门的映射唯一性
    dept_errors = {}
    for d in duties:
        if d.department:
            valid, err = _validate_duty_unique(d.department)
            if not valid:
                dept_errors[d.department] = err

    if dept_errors:
        unique_errors = list(set(dept_errors.values()))
        return jsonify({
            'error': '铁律一校验未通过，终止全部流程',
            'violations': unique_errors
        }), 400

    # 铁律三：按条线分组
    line_groups = {}
    for d in duties:
        if d.line_id:
            line_groups.setdefault(d.line_id, {'line': d.line, 'duties': [], 'catalogs': []})
            line_groups[d.line_id]['duties'].append(d)

    # 获取每个条线下的目录
    for line_id, group in line_groups.items():
        catalogs = Catalog.query.filter_by(line_id=line_id).all()
        group['catalogs'] = catalogs

    # 执行匹配
    all_results = []
    line_progress = []
    total_lines = len(line_groups)
    completed = 0

    for line_id, group in line_groups.items():
        line_name = group['line'].name if group['line'] else f'条线{line_id}'
        catalogs = group['catalogs']
        duties = group['duties']

        if not catalogs:
            line_progress.append({'line': line_name, 'status': '跳过', 'reason': '该条线下无目录数据'})
            continue

        # 构造入参
        catalog_items = [{'name': c.name, 'fields': json.loads(c.fields) if c.fields else []} for c in catalogs]
        duty_items = [{'name': d.duty_name, 'id': d.id} for d in duties]

        # 调用外部接口
        try:
            match_result = _call_external_api(catalog_items, duty_items)
        except Exception as e:
            line_progress.append({'line': line_name, 'status': '失败', 'reason': str(e)})
            continue

        # 保存匹配结果
        for c in catalogs:
            matched_duty_name = match_result.get(c.name, '')
            # 找到对应的 duty
            matched_duty = None
            for d in duties:
                if d.duty_name == matched_duty_name:
                    matched_duty = d
                    break

            result = MatchResult(
                catalog_id=c.id,
                duty_id=matched_duty.id if matched_duty else None,
                catalog_name=c.name,
                duty_name=matched_duty_name,
                department=duties[0].department if duties else '',
                line_name=line_name,
                matched_data=json.dumps(match_result, ensure_ascii=False)
            )
            db.session.add(result)
            all_results.append({
                'catalog_name': c.name,
                'duty_name': matched_duty_name,
                'department': duties[0].department if duties else '',
                'line_name': line_name
            })

        completed += 1
        line_progress.append({
            'line': line_name,
            'status': '成功',
            'catalog_count': len(catalogs),
            'duty_count': len(duties)
        })

    db.session.commit()

    return jsonify({
        'message': f'匹配完成，共处理 {total_lines} 个条线',
        'results_count': len(all_results),
        'results': all_results,
        'progress': line_progress,
        'completed_lines': completed,
        'total_lines': total_lines
    })


@duties_bp.route('/template')
def download_duty_template():
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '职能模板'
    ws.append(['部门名称', '内设机构', '职能名称'])
    ws.append(['例：财政局', '例：预算科', '例：负责全市财政预算编制'])
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='职能清单模板.xlsx')
